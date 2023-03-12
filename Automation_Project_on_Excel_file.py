# the problem being solved here is a pricing error in a spreadsheet that we want to always be able to fix
# by automatically running some code to generate it and return a corrected file

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
# print(sheet.max_row)

# used the above code section to identify how many rows are in the imported spreadsheet
# in order to make my for loop next

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')


# this code requires a fixed XLSX format in order to operate but it will correct the incorrect values everytime
# it should only be run with the prices are wrong and need to be fixed.
# if necessary you only need to change one number in the code to change the price correction amount
# i could have made it a price input required to further automate i might in the future update a commit fixing that



